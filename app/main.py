import json

from fastapi import Depends, FastAPI
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from sqlalchemy import func
from sqlalchemy.orm import Session

from app.database import get_db
from app.models import Customer, Email, Inquiry, Product, Quotation, QuotationItem
from app.services.ai_service import analyse
from datetime import datetime
from app.models import Customer, Draft, Email, Inquiry, PriceSetting, Product, Quotation
from app.services.gmail_service import create_draft
from app.services.draft_service import compose_quotation_reply
from app.services.draft_service import (compose_quotation_reply,
                                        compose_sample_reply,
                                        compose_delivery_reply,
                                        compose_after_sales_reply,
                                        compose_payment_reply,
                                        compose_blank_reply)

app = FastAPI(title="TradeQuote AI")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)


class EmailIn(BaseModel):
    subject: str
    body: str

LABELS = {"analysed": "Analysed", "quoted": "Quoted",
          "drafted": "Drafted", "sent": "Sent"}


def _status(e: Email) -> str:
    if not e.inquiry:
        return "Pending"
    return LABELS.get(e.inquiry.status, e.inquiry.status.title())


INCOTERM_KEYS = ("exw", "fca", "fob", "cfr", "cif", "cpt", "cip")


def _quote_prices(q: Quotation) -> dict:
    """把一筆 Quotation 的各 Incoterm 總價（整批貨，可能含多個品項）整理成
    compose_quotation_reply 要的格式。"""
    return {k: getattr(q, k) for k in INCOTERM_KEYS}


def _quote_items(q: Quotation) -> list:
    """把 Quotation 底下的品項轉成草稿模板要的格式，含每個品項的 EXW 單價。"""
    margin = q.margin_used or 0
    items = q.items or [QuotationItem(product=q.product, quantity=q.quantity,
                                       unit_price=q.product.unit_price)]
    return [{
        "product": item.product.name, "sku": item.product.sku,
        "quantity": item.quantity,
        "unit_exw": round(item.unit_price / (1 - margin), 4) if margin < 1 else 0,
        "moq": item.product.moq, "lead_days": item.product.lead_days,
    } for item in items]


@app.get("/")
def root():
    return {"service": "TradeQuote AI", "docs": "/docs"}


COUNTRIES = {"italy", "japan", "korea", "south korea", "china", "taiwan",
             "germany", "france", "spain", "egypt", "australia", "mexico",
             "brazil", "poland", "uae", "usa", "united states", "uk",
             "united kingdom", "india", "vietnam", "thailand", "singapore",
             "malaysia", "indonesia", "canada", "netherlands", "belgium",
             "turkey", "saudi arabia", "russia", "hong kong"}


def _clean(v):
    """模型偶爾會回字串 'null'，當成空值處理。"""
    if not v or not isinstance(v, str):
        return None
    s = v.strip()
    return None if s.lower() in {"null", "none", "n/a", "-", ""} else s


def _clean_company(v):
    """公司名不該是國家名。"""
    v = _clean(v)
    return None if v and v.lower() in COUNTRIES else v


def _resolve_customer(db, company, contact, sender_email, country):
    """依公司名 > 聯絡人 > 寄件信箱的優先順序找出或建立客戶。

    公司名沒擷取到時，用聯絡人姓名區分同一信箱底下的不同客戶——
    常見於測試信，或同一個信箱代轉多間公司詢價的情況，避免把明明不同的
    客戶都併成同一筆。真的兩者都沒有時才退回用寄件信箱本身當識別。
    """
    if company:
        cust = db.query(Customer).filter_by(company=company).first()
        if not cust:
            cust = Customer(company=company, contact=contact,
                             email=sender_email, country=country)
            db.add(cust)
            db.flush()
        elif contact and not cust.contact:
            cust.contact = contact
    elif contact:
        cust = db.query(Customer).filter_by(email=sender_email, contact=contact).first()
        if not cust:
            cust = Customer(company=contact, contact=contact,
                             email=sender_email, country=country)
            db.add(cust)
            db.flush()
    else:
        cust = db.query(Customer).filter_by(email=sender_email, contact=None).first()
        if not cust:
            cust = Customer(company="New Customer", email=sender_email, country=country)
            db.add(cust)
            db.flush()

    if country and not cust.country:
        cust.country = country
    return cust


def _display_company(i: Inquiry) -> str:
    """Inquiry 沒擷取到公司/聯絡人時的顯示名稱。

    other 意圖多半是廣告信、系統通知，根本不是客戶詢價，直接顯示寄件者本身
    讀到的名稱或信箱即可，不要顯示 New Customer——否則會跟真的詢價信但
    找不到署名的情況混在一起，分不出哪個是真的潛在客戶。
    """
    if i.company:
        return i.company
    if i.contact:
        return i.contact
    if i.intent == "other":
        return i.email.sender_name or i.email.sender_email
    return i.email.customer.company if i.email.customer else i.email.sender_email


def _freight_table(db):
    """從資料庫讀運費表，回傳 {國家: 費用}。"""
    from app.models import Freight
    return {f.destination: f.cost_usd for f in db.query(Freight).all()}


def _match_product(db, text):
    """完全符合名稱/別名優先，再來子字串比對，最後才是詞彙重疊。

    完全符合優先是必要的：例如查詢字串是「Charger」，Power Bank 的別名
    "portable charger" 用子字串比對也會命中（"charger" in "portable charger"），
    先做完全符合可以避免這種別名子字串誤配到別的商品。
    """
    if not text:
        return None
    n = text.lower().strip()
    products = db.query(Product).all()

    def names_of(p):
        names = [p.name.lower()]
        if p.aliases:
            names += [a.strip().lower() for a in p.aliases.split(",")]
        return names

    for p in products:
        if n in names_of(p):
            return p
    for p in products:
        if any(c in n or n in c for c in names_of(p)):
            return p

    words = {w.strip("-,.").rstrip("s") for w in n.split()}
    best, score = None, 0
    for p in products:
        target = {w.rstrip("s") for w in p.name.lower().split()}
        hits = len(words & target)
        if hits > score:
            best, score = p, hits
    return best if score else None


def _parse_items(inq: Inquiry) -> list:
    """把 Inquiry 存的 items_json 解析成 [{"product": str, "quantity": int|None}, ...]。
    舊資料沒有 items_json 時，退回用 product_text/quantity 包成一個品項。"""
    if inq.items_json:
        try:
            items = json.loads(inq.items_json)
            if items:
                return items
        except (json.JSONDecodeError, TypeError):
            pass
    if inq.product_text:
        return [{"product": inq.product_text, "quantity": inq.quantity}]
    return []


def _parse_incoterms(inq: Inquiry) -> list:
    """把 Inquiry 存的 incoterms_json 解析成大寫字串清單（客戶可能一次問好幾種
    條件，例如「FOB and CIF pricing」）。舊資料沒有 incoterms_json 時，退回用
    單一 incoterm 欄位包成一個元素的清單。"""
    if inq.incoterms_json:
        try:
            terms = json.loads(inq.incoterms_json)
            if terms:
                return terms
        except (json.JSONDecodeError, TypeError):
            pass
    return [inq.incoterm] if inq.incoterm else []


def _generate_quotation(db, inq: Inquiry, s, freight_lookup=None):
    """把一封詢價信裡所有品項各自比對商品、算成本，合成一張報價單——多個
    品項共用同一個 quote_no，運費/保險/出口地手續費只算一次（整批貨的費用，
    不會每個品項各分攤一次）。一個品項都比對不到商品就回傳 None。"""
    from app.services.pricing_service import calculate_from_cost

    matched = []
    for item in _parse_items(inq):
        p = _match_product(db, item.get("product"))
        if not p:
            continue
        qty = item.get("quantity") or p.moq
        matched.append((p, qty))

    if not matched:
        return None

    total_cost = sum(p.unit_price * qty for p, qty in matched)
    dest = inq.destination or ""
    if not dest:
        # 沒有目的地時：如果客戶要求的貿易條件全部都是「出貨港命名」
        # （EXW/FCA/FOB，運費本來就不影響這些條件的報價），改以我們自己
        # 的出貨港（高雄／台灣）為報價基準去查運費表；客戶完全沒指定
        # 條件、或條件裡有目的地命名的（CFR/CIF/CPT/CIP）就不能亂猜，
        # 維持空白，讓畫面提醒使用者確認目的地。
        origin_named = {"EXW", "FCA", "FOB"}
        reqs = {t.upper() for t in _parse_incoterms(inq)}
        if reqs and reqs.issubset(origin_named):
            dest = "Taiwan"
    calc = calculate_from_cost(total_cost, dest, s, freight_lookup)

    first_p, first_qty = matched[0]
    total_qty = sum(qty for _, qty in matched)
    n = db.query(Quotation).count() + 1
    q = Quotation(
        inquiry_id=inq.id, quote_no=f"Q-2026-{n:03d}",
        product_id=first_p.id, quantity=first_qty, destination=calc["destination"] or "",
        cost=calc["cost"], exw=calc["exw"], fca=calc["fca"], fob=calc["fob"],
        cfr=calc["cfr"], cif=calc["cif"], cpt=calc["cpt"], cip=calc["cip"],
        unit_cif=round(calc["cif"] / total_qty, 4) if total_qty else 0,
        margin_used=s.profit_margin, freight_used=calc["freight"],
        freight_estimated=calc["freight_estimated"],
    )
    db.add(q)
    db.flush()
    for p, qty in matched:
        db.add(QuotationItem(quotation_id=q.id, product_id=p.id, quantity=qty,
                              unit_price=p.unit_price, cost=p.unit_price * qty))
    return q


@app.get("/api/inbox")
def inbox(archived: bool = False, db: Session = Depends(get_db)):
    q = db.query(Email).filter(Email.ignored_at.is_(None))
    if archived:
        q = q.filter(Email.archived_at.isnot(None))
    else:
        q = q.filter(Email.archived_at.is_(None))
    rows = q.order_by(Email.id.desc()).all()
    return [{
        "message_id": e.message_id,
        "received": e.received_at,
        "company": (_display_company(e.inquiry) if e.inquiry
                    else (e.customer.company if e.customer else e.sender_email)),
        "email": e.sender_email,
        "subject": e.subject,
        "intent": e.inquiry.intent if e.inquiry else None,
        "confidence": e.inquiry.confidence if e.inquiry else None,
        "status": _status(e),
        "viewed": e.viewed_at is not None,
        
    } for e in rows]


@app.get("/api/inbox/{message_id}")
def email_detail(message_id: str, db: Session = Depends(get_db)):
    e = db.query(Email).filter_by(message_id=message_id).first()
    if not e:
        return {"error": "Not found"}

    out = {
        "email": {"subject": e.subject, "body": e.body,
                  "sender_name": e.sender_name, "sender_email": e.sender_email,
                  "received": e.received_at},
        "extracted": None, "quote": None, "draft": None,
    }
    if e.inquiry:
        i = e.inquiry
        out["extracted"] = {
            "intent": i.intent, "confidence": i.confidence,
            "company": i.company, "contact": i.contact,
            "product": i.product_text, "quantity": i.quantity,
            "items": _parse_items(i),
            "destination": i.destination, "incoterm": i.incoterm,
            "incoterms": _parse_incoterms(i),
            "summary": i.summary, "model": i.model_name,
        }
        if i.quotation:
            q = i.quotation
            settings_row = db.query(PriceSetting).first()
            out["quote"] = {
                "quote_no": q.quote_no, "sku": q.product.sku,
                "product": q.product.name, "quantity": q.quantity,
                "destination": q.destination, "freight_estimated": q.freight_estimated,
                "cost": q.cost, "exw": q.exw, "fca": q.fca, "fob": q.fob,
                "cfr": q.cfr, "cif": q.cif, "cpt": q.cpt, "cip": q.cip,
                "unit_cif": q.unit_cif,
                "margin": q.margin_used, "freight": q.freight_used,
                "moq": q.product.moq, "lead_days": q.product.lead_days,
                "items": [{"sku": it.product.sku, "product": it.product.name,
                           "quantity": it.quantity, "unit_price": it.unit_price,
                           "cost": it.cost} for it in q.items],
                "breakdown": {
                    "cost": q.cost,
                    "margin_pct": q.margin_used,
                    "margin_amount": round(q.cost * q.margin_used / (1 - q.margin_used), 2) if q.margin_used < 1 else 0,
                    "freight": q.freight_used,
                    "local_charges": settings_row.local_charges,
                    "insurance": settings_row.insurance,
                    "bank_charges": settings_row.bank_charges,
                },
            }

        from app.models import Draft
        d = db.query(Draft).filter_by(inquiry_id=i.id).first()
        if d:
            out["draft"] = {
                "id": d.id, "to": d.to_email, "subject": d.subject,
                "body": d.body, "status": d.status,
                "gmail_draft_id": d.gmail_draft_id,
            }
    return out


@app.get("/api/stats")
def stats(db: Session = Depends(get_db)):
    dist = dict(db.query(Inquiry.intent, func.count())
                  .group_by(Inquiry.intent).all())
    from app.models import Draft
    return {
        "emails": db.query(Email).filter(Email.ignored_at.is_(None)).count(),
        "analysed": db.query(Inquiry).count(),
        "quotations": db.query(Quotation).count(),
        "customers": db.query(Customer.id).join(Email, Email.customer_id == Customer.id)
                       .filter(Email.ignored_at.is_(None)).distinct().count(),
        "unread": db.query(Email).filter(
            Email.viewed_at.is_(None), Email.archived_at.is_(None),
            Email.ignored_at.is_(None)).count(),
        "pending_drafts": db.query(Draft).filter(Draft.status == "draft").count(),
        "intent_distribution": dist,
    }


@app.get("/api/products")
def products(db: Session = Depends(get_db)):
    return [{"sku": p.sku, "name": p.name, "price": p.unit_price,
             "moq": p.moq, "lead": p.lead_days, "hs_code": p.hs_code}
            for p in db.query(Product).order_by(Product.sku).all()]


@app.get("/api/quotations")
def quotations(db: Session = Depends(get_db)):
    settings_row = db.query(PriceSetting).first()
    rows = []
    for q in db.query(Quotation).order_by(Quotation.id.desc()).all():
        i = q.inquiry
        draft = db.query(Draft).filter_by(quotation_id=q.id).first()
        rows.append({
            "quote_no": q.quote_no, "company": _display_company(i),
            "product": q.product.name, "quantity": q.quantity,
            "incoterms": _parse_incoterms(i),
            "items": [{"sku": it.product.sku, "product": it.product.name,
                       "quantity": it.quantity, "unit_price": it.unit_price,
                       "cost": it.cost} for it in q.items],
            "destination": q.destination, "freight_estimated": q.freight_estimated,
            "exw": q.exw, "fca": q.fca, "fob": q.fob, "cfr": q.cfr,
            "cif": q.cif, "cpt": q.cpt, "cip": q.cip, "status": q.status,
            "draft": {
                "id": draft.id,
                "outdated": q.updated_at > draft.updated_at,
            } if draft else None,
            "breakdown": {
                "cost": q.cost,
                "margin_pct": q.margin_used,
                "margin_amount": round(q.cost * q.margin_used / (1 - q.margin_used), 2) if q.margin_used < 1 else 0,
                "freight": q.freight_used,
                "local_charges": settings_row.local_charges,
                "insurance": settings_row.insurance,
                "bank_charges": settings_row.bank_charges,
            },
            "extraction": {
                "confidence": i.confidence, "company": i.company, "contact": i.contact,
                "destination": i.destination, "incoterms": _parse_incoterms(i),
                "summary": i.summary, "model": i.model_name,
                "received": i.email.received_at, "sender_email": i.email.sender_email,
                "subject": i.email.subject,
                "items": _parse_items(i),
            },
        })
    return rows


@app.get("/api/customers")
def customers(db: Session = Depends(get_db)):
    rows = []
    for c in db.query(Customer).all():
        active = [e for e in c.emails if e.ignored_at is None]
        if not active:
            continue
        quotes = [e.inquiry.quotation for e in active
                  if e.inquiry and e.inquiry.quotation]
        last = max(active, key=lambda e: e.fetched_at)
        rows.append({
            "id": c.id, "company": c.company, "email": c.email,
            "country": c.country, "language": c.language,
            "industry": c.industry, "emails": len(active),
            "quotations": len(quotes),
            "total_value": sum(q.cif for q in quotes),
            "last_contact": last.received_at,
        })
    return sorted(rows, key=lambda r: -r["emails"])


@app.get("/api/customers/{customer_id}")
def customer_detail(customer_id: int, db: Session = Depends(get_db)):
    c = db.query(Customer).filter_by(id=customer_id).first()
    if not c:
        return {"error": "Not found"}

    active = [e for e in c.emails if e.ignored_at is None]
    history = []
    for e in sorted(active, key=lambda e: e.fetched_at, reverse=True):
        i = e.inquiry
        q = i.quotation if i else None
        history.append({
            "message_id": e.message_id,
            "subject": e.subject,
            "received": e.received_at,
            "intent": i.intent if i else None,
            "status": _status(e),
            "quote_no": q.quote_no if q else None,
            "product": (", ".join(it.product.name for it in q.items) if q and q.items
                        else (q.product.name if q else (i.product_text if i else None))),
            "quantity": (sum(it.quantity for it in q.items) if q and q.items
                        else (q.quantity if q else (i.quantity if i else None))),
            "destination": q.destination if q else (i.destination if i else None),
            "cif": q.cif if q else None,
        })

    return {
        "id": c.id, "company": c.company, "email": c.email,
        "contact": c.contact, "country": c.country, "language": c.language,
        "industry": c.industry,
        "quotations": sum(1 for h in history if h["quote_no"]),
        "total_value": sum(h["cif"] for h in history if h["cif"]),
        "history": history,
    }


@app.post("/api/analyse")
def analyse_email(payload: EmailIn):
    try:
        return analyse(payload.subject, payload.body)
    except Exception as e:
        return {"error": str(e)}
    
@app.post("/api/inbox/{message_id}/viewed")
def mark_viewed(message_id: str, db: Session = Depends(get_db)):
    e = db.query(Email).filter_by(message_id=message_id).first()
    if e and not e.viewed_at:
        e.viewed_at = datetime.now()
        db.commit()
    return {"ok": True}

class SettingsIn(BaseModel):
    profit_margin: float | None = None
    local_charges: float | None = None
    insurance: float | None = None
    bank_charges: float | None = None
    usd_twd: float | None = None
    sync_limit: int | None = None
    shipping_port: str | None = None


@app.get("/api/settings")
def get_settings(db: Session = Depends(get_db)):
    s = db.query(PriceSetting).first()
    if not s:
        s = PriceSetting()
        db.add(s)
        db.commit()
    return {
        "profit_margin": s.profit_margin,
        "local_charges": s.local_charges,
        "insurance": s.insurance,
        "bank_charges": s.bank_charges,
        "usd_twd": s.usd_twd,
        "sync_limit": s.sync_limit,
        "shipping_port": s.shipping_port,
        "updated_at": s.updated_at.strftime("%Y-%m-%d %H:%M"),
    }


@app.put("/api/settings")
def update_settings(payload: SettingsIn, db: Session = Depends(get_db)):
    s = db.query(PriceSetting).first()
    if not s:
        s = PriceSetting()
        db.add(s)
    for field, value in payload.model_dump().items():
        if value is not None:
            setattr(s, field, value)
    db.commit()
    return {"ok": True, "updated_at": s.updated_at.strftime("%Y-%m-%d %H:%M")}


@app.get("/api/freight")
def freight(db: Session = Depends(get_db)):
    from app.models import Freight
    return [{"destination": f.destination, "port": f.port,
             "cost_usd": f.cost_usd, "transit_days": f.transit_days}
            for f in db.query(Freight).order_by(Freight.destination).all()]


class FreightIn(BaseModel):
    destination: str
    port: str | None = None
    cost_usd: float
    transit_days: int | None = None


@app.post("/api/freight")
def create_freight(payload: FreightIn, db: Session = Depends(get_db)):
    from app.models import Freight
    destination = payload.destination.strip()
    if not destination:
        return {"error": "Destination is required"}
    if db.query(Freight).filter_by(destination=destination).first():
        return {"error": f"{destination} already exists"}
    f = Freight(destination=destination, port=payload.port,
                cost_usd=payload.cost_usd, transit_days=payload.transit_days)
    db.add(f)
    db.commit()
    return {"ok": True, "destination": f.destination}


@app.delete("/api/freight/{destination}")
def delete_freight(destination: str, db: Session = Depends(get_db)):
    from app.models import Freight
    f = db.query(Freight).filter_by(destination=destination).first()
    if not f:
        return {"error": "Not found"}
    db.delete(f)
    db.commit()
    return {"ok": True}


@app.get("/api/drafts")
def list_drafts(db: Session = Depends(get_db)):
    rows = db.query(Draft).order_by(Draft.id.desc()).all()
    return [{
        "id": d.id,
        "to": d.to_email,
        "subject": d.subject,
        "body": d.body,
        "status": d.status,
        "company": _display_company(d.inquiry),
        "gmail_draft_id": d.gmail_draft_id,
        "updated_at": d.updated_at.strftime("%Y-%m-%d %H:%M"),
    } for d in rows]


@app.post("/api/quotations/{quote_no}/draft")
def generate_draft(quote_no: str, db: Session = Depends(get_db)):
    """從報價產生一封草稿，存進資料庫（還沒送 Gmail）。"""
    q = db.query(Quotation).filter_by(quote_no=quote_no).first()
    if not q:
        return {"error": "Quotation not found"}

    existing = db.query(Draft).filter_by(inquiry_id=q.inquiry_id).first()
    if existing:
        return {"error": "Draft already exists", "draft_id": existing.id}

    inq = q.inquiry
    email = inq.email
    settings_row = db.query(PriceSetting).first()
    body = compose_quotation_reply(
        contact=inq.contact, items=_quote_items(q), dest=q.destination,
        prices=_quote_prices(q), incoterms=_parse_incoterms(inq),
        shipping_port=settings_row.shipping_port,
    )
    d = Draft(
        quotation_id=q.id, inquiry_id=inq.id,
        to_email=email.sender_email,
        subject="Re: " + (email.subject or "Your inquiry"),
        body=body,
    )
    db.add(d)
    inq.status = "drafted"
    q.status = "drafted"
    db.commit()
    return {"ok": True, "draft_id": d.id}


class DraftEdit(BaseModel):
    subject: str
    body: str


@app.put("/api/drafts/{draft_id}")
def edit_draft(draft_id: int, payload: DraftEdit, db: Session = Depends(get_db)):
    d = db.query(Draft).filter_by(id=draft_id).first()
    if not d:
        return {"error": "Not found"}
    d.subject = payload.subject
    d.body = payload.body
    db.commit()
    return {"ok": True}


@app.post("/api/drafts/{draft_id}/send-to-gmail")
def send_to_gmail(draft_id: int, db: Session = Depends(get_db)):
    """在 Gmail 建立草稿（不是寄出，是存進草稿匣）。"""
    d = db.query(Draft).filter_by(id=draft_id).first()
    if not d:
        return {"error": "Not found"}
    try:
        gmail_id = create_draft(d.to_email, d.subject, d.body)
    except Exception as e:
        return {"error": "Gmail failed: " + str(e)}
    d.gmail_draft_id = gmail_id
    d.status = "sent"
    if d.inquiry.quotation:
        d.inquiry.quotation.draft_id = gmail_id
        d.inquiry.quotation.status = "sent"
    d.inquiry.status = "sent"
    if d.inquiry.email:
        d.inquiry.email.archived_at = datetime.now()
    db.commit()
    return {"ok": True, "gmail_draft_id": gmail_id}


@app.post("/api/quotations/{quote_no}/recalculate")
def recalculate(quote_no: str, db: Session = Depends(get_db)):
    """用目前的 Price Settings 重算,更新快照（含每個品項用目前的產品單價重算）。"""
    from app.services.pricing_service import PriceSettings, calculate_from_cost

    q = db.query(Quotation).filter_by(quote_no=quote_no).first()
    if not q:
        return {"error": "Not found"}

    row = db.query(PriceSetting).first()
    s = PriceSettings(
        profit_margin=row.profit_margin,
        local_charges=row.local_charges,
        insurance=row.insurance,
        bank_charges=row.bank_charges,
        usd_twd=row.usd_twd,
    )
    total_cost = 0.0
    total_qty = 0
    for it in q.items:
        it.unit_price = it.product.unit_price
        it.cost = it.unit_price * it.quantity
        total_cost += it.cost
        total_qty += it.quantity

    dest = q.destination or ""
    if not dest:
        origin_named = {"EXW", "FCA", "FOB"}
        reqs = {t.upper() for t in _parse_incoterms(q.inquiry)}
        if reqs and reqs.issubset(origin_named):
            dest = "Taiwan"
    calc = calculate_from_cost(total_cost, dest, s, _freight_table(db))
    q.destination = calc["destination"] or ""
    q.cost = calc["cost"]
    q.exw = calc["exw"]
    q.fca = calc["fca"]
    q.fob = calc["fob"]
    q.cfr = calc["cfr"]
    q.cif = calc["cif"]
    q.cpt = calc["cpt"]
    q.cip = calc["cip"]
    q.unit_cif = round(calc["cif"] / total_qty, 4) if total_qty else 0
    q.margin_used = s.profit_margin
    q.freight_used = calc["freight"]
    q.freight_estimated = calc["freight_estimated"]
    db.commit()
    return {"ok": True, "cif": q.cif, "margin": q.margin_used}

@app.post("/api/inbox/sync")
def sync_inbox(query: str = "is:unread", limit: int | None = None,
               db: Session = Depends(get_db)):
    """抓 Gmail 新信 → AI 分析 → 算價 → 入庫。回傳新增數量。"""
    from app.services.ai_service import MODEL, analyse
    from app.services.gmail_service import fetch_new
    from app.services.pricing_service import PriceSettings

    row = db.query(PriceSetting).first()
    s = PriceSettings(
        profit_margin=row.profit_margin, local_charges=row.local_charges,
        insurance=row.insurance, bank_charges=row.bank_charges, usd_twd=row.usd_twd,
    )
    limit = limit or row.sync_limit
    freight_lookup = _freight_table(db)

    new = 0
    try:
        messages = fetch_new(limit, query)
    except Exception as e:
        return {"error": "Gmail failed: " + str(e)}

    for m in messages:
        existing = db.query(Email).filter_by(message_id=m["message_id"]).first()
        if existing:
            continue

        email = Email(
            message_id=m["message_id"], sender_name=m["sender_name"],
            sender_email=m["sender_email"], subject=m["subject"],
            body=m["body"], received_at=m["date"],
        )
        db.add(email)
        db.flush()

        try:
            r = analyse(m["subject"], m["body"])
        except Exception:
            db.commit()
            continue

        company = _clean_company(r.get("company"))
        contact = _clean(r.get("contact"))
        country = _clean(r.get("destination"))
        cust = _resolve_customer(db, company, contact, m["sender_email"], country)
        email.customer_id = cust.id

        inq = Inquiry(
            email_id=email.id,
            intent=(r.get("intent") if r.get("intent") in
                    {"quotation", "sample_request", "delivery_followup",
                     "after_sales", "payment", "other"} else "other"),
            confidence=r.get("confidence", 0), company=r.get("company"),
            contact=r.get("contact"), product_text=r.get("product"),
            quantity=r.get("quantity"), items_json=json.dumps(r.get("items") or []),
            destination=r.get("destination"),
            incoterm=r.get("incoterm"), incoterms_json=json.dumps(r.get("incoterms") or []),
            summary=r.get("summary"), model_name=MODEL,
        )
        db.add(inq)
        db.flush()

        if inq.intent == "quotation" and _generate_quotation(db, inq, s, freight_lookup):
            inq.status = "quoted"

        db.commit()
        new += 1

    return {"ok": True, "new": new}

@app.get("/api/reports")
def reports(db: Session = Depends(get_db)):
    from sqlalchemy import func

    total_emails = db.query(Email).filter(Email.ignored_at.is_(None)).count()
    total_quotes = db.query(Quotation).count()

    # intent 分布
    dist = dict(db.query(Inquiry.intent, func.count())
                  .group_by(Inquiry.intent).all())

    # 各產品被報價次數 + 總金額
    prod_rows = (db.query(Product.name,
                          func.count(Quotation.id),
                          func.sum(Quotation.cif))
                   .join(Quotation, Quotation.product_id == Product.id)
                   .group_by(Product.name)
                   .order_by(func.count(Quotation.id).desc())
                   .all())
    products = [{"name": n, "count": c, "value": float(v or 0)}
                for n, c, v in prod_rows]

    # 報價狀態分布(成交率用)
    status_rows = dict(db.query(Quotation.status, func.count())
                         .group_by(Quotation.status).all())

    # 目的地分布
    dest_rows = (db.query(Quotation.destination, func.count())
                   .group_by(Quotation.destination)
                   .order_by(func.count(Quotation.id).desc()).all())
    destinations = [{"name": d, "count": c} for d, c in dest_rows]

    total_value = db.query(func.sum(Quotation.cif)).scalar() or 0
    avg_conf = db.query(func.avg(Inquiry.confidence)).scalar() or 0

    return {
        "total_emails": total_emails,
        "total_quotes": total_quotes,
        "total_value": float(total_value),
        "avg_confidence": round(float(avg_conf)),
        "intent_distribution": dist,
        "top_products": products,
        "destinations": destinations,
        "quote_status": status_rows,
    }
    
@app.get("/api/system-info")
def system_info(db: Session = Depends(get_db)):
    import os
    from app.services.ai_service import MODEL

    # 檢查 Ollama 是否在線
    ollama_ok = False
    try:
        import httpx
        r = httpx.get("http://localhost:11434/api/tags", timeout=2)
        ollama_ok = r.status_code == 200
    except Exception:
        ollama_ok = False

    # 從 token.json 判斷 Gmail 是否授權
    gmail_ok = os.path.exists("token.json")

    return {
        "model": MODEL,
        "model_runtime": "Ollama (local)",
        "ollama_online": ollama_ok,
        "gmail_authorised": gmail_ok,
        "data_location": "Local SQLite (tradequote.db)",
        "privacy_note": "All email analysis runs on this machine. No data leaves the device.",
    }


@app.post("/api/system/reset")
def reset_database():
    """清空所有資料表並重新灌入預設資料，等同手動執行
    `rm tradequote.db && uv run python seed.py`，但不會動到 Gmail 憑證。"""
    from app.database import SessionLocal, engine
    from app.models import Base
    from seed import seed as run_seed

    Base.metadata.drop_all(engine)
    Base.metadata.create_all(engine)
    db = SessionLocal()
    try:
        run_seed(db)
    finally:
        db.close()
    return {"ok": True}


@app.get("/api/export")
def export_excel(db: Session = Depends(get_db)):
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from fastapi.responses import StreamingResponse

    wb = Workbook()
    wb.remove(wb.active)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1E293B")

    def write_sheet(title, headers, rows):
        ws = wb.create_sheet(title[:31])
        ws.append(headers)
        for c in ws[1]:
            c.font = header_font
            c.fill = header_fill
        for r in rows:
            ws.append(r)
        for i, h in enumerate(headers, 1):
            width = max(len(str(h)), *(len(str(r[i-1])) for r in rows)) if rows else len(h)
            ws.column_dimensions[chr(64+i)].width = min(width + 4, 50)

    intents = ["quotation", "sample_request", "delivery_followup", "after_sales", "payment"]
    for intent in intents:
        rows = db.query(Inquiry).filter_by(intent=intent).all()
        write_sheet(
            intent.replace("_", " ").title(),
            ["Date", "Company", "Contact", "Product", "Qty", "Destination", "Confidence", "Summary"],
            [[i.analysed_at.strftime("%Y-%m-%d %H:%M"), i.company or "", i.contact or "",
              i.product_text or "", i.quantity or "", i.destination or "",
              f"{i.confidence}%", i.summary or ""] for i in rows],
        )

    quotes = db.query(Quotation).all()
    write_sheet(
        "Quotations",
        ["Quote No", "Company", "Product", "Qty", "Destination", "EXW", "FOB", "CIF", "Status"],
        [[q.quote_no, q.inquiry.company or "", q.product.name, q.quantity, q.destination,
          round(q.exw), round(q.fob), round(q.cif), q.status] for q in quotes],
    )

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=tradequote_export.xlsx"},
    )
    
@app.post("/api/inbox/{message_id}/quote")
def quote_from_email(message_id: str, db: Session = Depends(get_db)):
    """對這封信產生報價。已有報價就直接回傳。"""
    from app.services.pricing_service import PriceSettings

    e = db.query(Email).filter_by(message_id=message_id).first()
    if not e or not e.inquiry:
        return {"error": "Email not analysed yet"}

    i = e.inquiry
    if i.quotation:
        return {"ok": True, "quote_no": i.quotation.quote_no, "existing": True}

    if i.intent != "quotation":
        return {"error": "This email is not a quotation request"}

    row = db.query(PriceSetting).first()
    s = PriceSettings(
        profit_margin=row.profit_margin, local_charges=row.local_charges,
        insurance=row.insurance, bank_charges=row.bank_charges, usd_twd=row.usd_twd,
    )
    q = _generate_quotation(db, i, s, _freight_table(db))
    if not q:
        return {"error": f"No product in the catalogue matches '{i.product_text}'"}

    i.status = "quoted"
    db.commit()
    return {"ok": True, "quote_no": q.quote_no}


@app.post("/api/inbox/{message_id}/draft")
def draft_from_email(message_id: str, db: Session = Depends(get_db)):
    """對這封信的報價產生草稿。"""
    from app.models import Draft

    e = db.query(Email).filter_by(message_id=message_id).first()
    if not e or not e.inquiry:
        return {"error": "Email not analysed yet"}

    i = e.inquiry
    existing = db.query(Draft).filter_by(inquiry_id=i.id).first()
    if existing:
        return {"ok": True, "draft_id": existing.id, "existing": True}

    q = i.quotation
    if not q:
        return {"error": "Generate a quotation first"}

    settings_row = db.query(PriceSetting).first()
    body = compose_quotation_reply(
        contact=i.contact, items=_quote_items(q), dest=q.destination,
        prices=_quote_prices(q), incoterms=_parse_incoterms(i),
        shipping_port=settings_row.shipping_port,
    )
    d = Draft(
        quotation_id=q.id, inquiry_id=i.id,
        to_email=e.sender_email,
        subject="Re: " + (e.subject or "Your inquiry"),
        body=body,
    )
    db.add(d)
    i.status = "drafted"
    q.status = "drafted"
    db.commit()
    return {"ok": True, "draft_id": d.id}
@app.post("/api/drafts/{draft_id}/regenerate")
def regenerate_draft(draft_id: int, db: Session = Depends(get_db)):
    """用目前的報價重新產生草稿內容（覆蓋原本的）。"""
    from app.models import Draft

    d = db.query(Draft).filter_by(id=draft_id).first()
    if not d:
        return {"error": "Not found"}

    i = d.inquiry
    q = i.quotation
    if not q:
        return {"error": "No quotation attached to this draft"}

    settings_row = db.query(PriceSetting).first()
    d.body = compose_quotation_reply(
        contact=i.contact, items=_quote_items(q), dest=q.destination,
        prices=_quote_prices(q), incoterms=_parse_incoterms(i),
        shipping_port=settings_row.shipping_port,
    )
    db.commit()
    return {"ok": True}

TEMPLATES = {
    "sample_request": lambda i: compose_sample_reply(i.contact, i.product_text, i.quantity),
    "delivery_followup": lambda i: compose_delivery_reply(i.contact, None),
    "after_sales": lambda i: compose_after_sales_reply(i.contact, i.product_text, i.quantity),
    "payment": lambda i: compose_payment_reply(i.contact, None),
}


@app.post("/api/inbox/{message_id}/reply")
def reply_from_email(message_id: str, mode: str = "template",
                     db: Session = Depends(get_db)):
    """為非報價信產生草稿。mode = template | blank"""
    from app.models import Draft

    e = db.query(Email).filter_by(message_id=message_id).first()
    if not e or not e.inquiry:
        return {"error": "Email not analysed yet"}

    i = e.inquiry
    existing = db.query(Draft).filter_by(inquiry_id=i.id).first()
    if existing:
        return {"ok": True, "draft_id": existing.id, "existing": True}

    if mode == "blank":
        body = compose_blank_reply(i.contact)
    else:
        maker = TEMPLATES.get(i.intent)
        body = maker(i) if maker else compose_blank_reply(i.contact)

    d = Draft(
        inquiry_id=i.id,
        to_email=e.sender_email,
        subject="Re: " + (e.subject or "Your enquiry"),
        body=body,
    )
    db.add(d)
    i.status = "drafted"
    db.commit()
    return {"ok": True, "draft_id": d.id}


@app.post("/api/inbox/{message_id}/archive")
def archive_email(message_id: str, db: Session = Depends(get_db)):
    e = db.query(Email).filter_by(message_id=message_id).first()
    if not e:
        return {"error": "Not found"}
    e.archived_at = datetime.now()
    db.commit()
    return {"ok": True}


@app.post("/api/inbox/{message_id}/unarchive")
def unarchive_email(message_id: str, db: Session = Depends(get_db)):
    e = db.query(Email).filter_by(message_id=message_id).first()
    if not e:
        return {"error": "Not found"}
    e.archived_at = None
    db.commit()
    return {"ok": True}


@app.post("/api/inbox/{message_id}/delete")
def delete_email(message_id: str, db: Session = Depends(get_db)):
    """永久刪除一封信：清內容、保留 message_id 避免重新同步時又抓回來。

    常見情境：Irrelevant 頁籤誤按 Keep（信件因此移到 Active），這裡讓使用者
    在 Active 頁籤補救刪除。跟 purge_irrelevant 用一樣的「清內容留 message_id」
    做法，但這裡不限定 intent，任何一封 Active 的信都能刪。
    """
    from app.models import Draft

    e = db.query(Email).filter_by(message_id=message_id).first()
    if not e:
        return {"error": "Not found"}

    if e.inquiry:
        d = db.query(Draft).filter_by(inquiry_id=e.inquiry.id).first()
        if d:
            db.delete(d)
        if e.inquiry.quotation:
            db.delete(e.inquiry.quotation)
        db.delete(e.inquiry)

    e.body = ""
    e.subject = e.subject or ""
    e.ignored_at = datetime.now()
    db.commit()
    return {"ok": True}
@app.get("/api/irrelevant")
def irrelevant_emails(db: Session = Depends(get_db)):
    rows = (db.query(Email).join(Inquiry)
              .filter(Inquiry.intent == "other", Email.ignored_at.is_(None),
                      Inquiry.reviewed.is_(False))
              .order_by(Email.id.desc()).all())
    return [{
        "message_id": e.message_id,
        "received": e.received_at,
        "company": (e.inquiry.company if e.inquiry and e.inquiry.company
                    else e.sender_name),
        "email": e.sender_email,
        "subject": e.subject,
        "summary": e.inquiry.summary if e.inquiry else None,
        "confidence": e.inquiry.confidence if e.inquiry else None,
    } for e in rows]


@app.post("/api/irrelevant/purge")
def purge_irrelevant(db: Session = Depends(get_db)):
    """清除所有不相關的信：刪掉內容，保留 message_id 供去重。"""
    rows = (db.query(Email).join(Inquiry)
              .filter(Inquiry.intent == "other", Email.ignored_at.is_(None),
                      Inquiry.reviewed.is_(False)).all())
    n = 0
    for e in rows:
        if e.inquiry:
            db.delete(e.inquiry)
        e.body = ""
        e.subject = e.subject or ""
        e.ignored_at = datetime.now()
        n += 1
    db.commit()
    return {"ok": True, "purged": n}


@app.post("/api/inbox/{message_id}/keep")
def keep_email(message_id: str, db: Session = Depends(get_db)):
    """把誤判的信留下來：產生空白草稿，讓使用者自己寫。"""
    from app.models import Draft

    e = db.query(Email).filter_by(message_id=message_id).first()
    if not e or not e.inquiry:
        return {"error": "Not found"}

    i = e.inquiry
    i.reviewed = True

    existing = db.query(Draft).filter_by(inquiry_id=i.id).first()
    if existing:
        db.commit()
        return {"ok": True, "draft_id": existing.id}

    d = Draft(
        inquiry_id=i.id,
        to_email=e.sender_email,
        subject="Re: " + (e.subject or "Your enquiry"),
        body=compose_blank_reply(i.contact),
    )
    db.add(d)
    i.status = "drafted"
    db.commit()
    return {"ok": True, "draft_id": d.id}

class ProductIn(BaseModel):
    sku: str
    name: str
    unit_price: float
    moq: int
    lead_days: int
    weight_kg: float | None = None
    hs_code: str | None = None
    aliases: str | None = None


@app.post("/api/products")
def create_product(payload: ProductIn, db: Session = Depends(get_db)):
    if db.query(Product).filter_by(sku=payload.sku).first():
        return {"error": f"SKU {payload.sku} already exists"}
    p = Product(**payload.model_dump())
    db.add(p)
    db.commit()
    return {"ok": True, "sku": p.sku}


@app.put("/api/products/{sku}")
def update_product(sku: str, payload: ProductIn, db: Session = Depends(get_db)):
    p = db.query(Product).filter_by(sku=sku).first()
    if not p:
        return {"error": "Not found"}
    for k, v in payload.model_dump().items():
        setattr(p, k, v)
    db.commit()
    return {"ok": True}


@app.delete("/api/products/{sku}")
def delete_product(sku: str, db: Session = Depends(get_db)):
    p = db.query(Product).filter_by(sku=sku).first()
    if not p:
        return {"error": "Not found"}
    if db.query(Quotation).filter_by(product_id=p.id).first():
        return {"error": "This product has quotations and cannot be deleted"}
    db.delete(p)
    db.commit()
    return {"ok": True}